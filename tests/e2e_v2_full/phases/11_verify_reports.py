#!/usr/bin/env python3
"""
Phase 11 — Verify reports artifacts + GitHub sandbox state.

Read-only verification of the artifacts produced by Wave B Units 8-10:
- Per-person Markdown timesheets + team rollup
- Per-iteration JSON results + frozen snapshots
- XLSX exports (parsed via openpyxl)
- PI-level summaries (pi_close + reports.py --pi)
- GitHub sandbox repo state (merged PRs, CI workflow runs, archive flag)

Phase contract
--------------
- Sandbox dir resolution order:
    1. ``EDPA_E2E_SANDBOX_DIR`` env var (set by the run_e2e.sh harness)
    2. Fallback to the Wave B run sandbox path (``/tmp/edpa-e2e-20260527-142316-c6ac4db8``).
       The run is identified by the ``EDPA_E2E_RUN_TAG`` baked into the path.
- GitHub repo resolution order:
    1. ``EDPA_E2E_GH_REPO`` env var (``owner/repo`` form)
    2. Fallback derived from ``EDPA_E2E_GH_OWNER``/``EDPA_E2E_RUN_TAG``
    3. Fallback to ``technomaton/edpa-e2e-20260527-142316-c6ac4db8`` (Wave B sandbox).
- Exit codes:
    * 0 — every check passed
    * 1 — one or more checks failed (details printed to stdout + stderr)
    * 2 — preflight error (sandbox dir missing, gh missing, etc.)
- The script never mutates the sandbox or the GitHub repo.

Expected artifacts (from Wave B Unit 10 run log)
------------------------------------------------
- 10 iteration report dirs under ``.edpa/reports/iteration-PI-2026-{1,2}.{1..5}/``
- Each iteration dir contains:
    * ``edpa_results.json``
    * ``edpa-results.xlsx``
    * 5 × ``timesheet-<person>.md`` (alice, bob-arch, bob-pm, carol, dave)
    * ``timesheet-team.md``
- 2 PI summary dirs under ``.edpa/reports/pi-PI-2026-{1,2}/`` containing
  ``pi_results.json``, ``summary.md`` (from ``pi_close.py``) and
  ``pi-summary-PI-2026-{1,2}.md`` (from ``reports.py --pi``).
- 12 frozen snapshots under ``.edpa/snapshots/``
  (10 base + 2 revisions from PI-2026-1.1 discovery).
- 24 merged PRs on GitHub, 24 ``edpa-contribution-sync.yml`` workflow runs
  (23 success + 1 acknowledged failure on S-5).
"""

from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

DEFAULT_SANDBOX = Path("/tmp/edpa-e2e-20260527-142316-c6ac4db8")
DEFAULT_REPO = "technomaton/edpa-e2e-20260527-142316-c6ac4db8"

# People IDs the engine writes one timesheet per (see Wave B Unit 10 log).
PEOPLE = ("alice", "bob-arch", "bob-pm", "carol", "dave")

# 10 iterations spread across two PIs.
ITERATIONS = tuple(
    f"PI-2026-{pi}.{n}" for pi in (1, 2) for n in (1, 2, 3, 4, 5)
)
PI_IDS = ("PI-2026-1", "PI-2026-2")

# IP iterations: events EV-1/EV-2 have no .md backlog entry, so the engine
# materialises zero derived hours for those iterations. Documented in
# Wave B Unit 10 run log as expected.
IP_ITERATIONS = {"PI-2026-1.5", "PI-2026-2.5"}

# Allow either old (vykaz-) or new (timesheet-) naming convention so the
# script does not break if a future patch renames the files.
TIMESHEET_PREFIXES = ("timesheet-", "vykaz-")
JSON_SIDECAR_PREFIXES = ("timesheet-", "vykaz-")

EXPECTED_MERGED_PRS = 24       # 14 PI-1 + 10 PI-2
EXPECTED_CI_RUNS_MIN = 14      # PI-1 minimum (real CI); hybrid may add more


@dataclass
class IterationReport:
    iteration_id: str
    md_count: int = 0
    json_sidecar_count: int = 0
    has_xlsx: bool = False
    xlsx_sheets: tuple[str, ...] = ()
    has_results_json: bool = False
    team_total: float | None = None
    invariants_passed: bool | None = None
    issues: list[str] = field(default_factory=list)


@dataclass
class VerifyResult:
    iterations: list[IterationReport] = field(default_factory=list)
    pi_summaries: dict[str, list[str]] = field(default_factory=dict)
    snapshot_count: int = 0
    merged_pr_count: int | None = None
    ci_run_counts: dict[str | None, int] = field(default_factory=dict)
    repo_archived: bool | None = None
    failures: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return not self.failures


# -----------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------

def _resolve_sandbox() -> Path:
    raw = os.environ.get("EDPA_E2E_SANDBOX_DIR")
    if raw:
        return Path(raw)
    return DEFAULT_SANDBOX


def _resolve_repo() -> str:
    explicit = os.environ.get("EDPA_E2E_GH_REPO")
    if explicit:
        return explicit
    owner = os.environ.get("EDPA_E2E_GH_OWNER")
    tag = os.environ.get("EDPA_E2E_RUN_TAG")
    if owner and tag:
        return f"{owner}/edpa-e2e-{tag}"
    return DEFAULT_REPO


def _iter_timesheets(iter_dir: Path, suffix: str) -> list[Path]:
    """Return per-person timesheets (excluding the team rollup) with given suffix."""
    files: list[Path] = []
    for prefix in TIMESHEET_PREFIXES:
        for path in iter_dir.glob(f"{prefix}*{suffix}"):
            stem = path.stem  # e.g. "timesheet-alice"
            if stem.endswith("-team"):
                continue
            files.append(path)
    return sorted(files)


def _person_id_from_filename(path: Path) -> str:
    name = path.stem
    for prefix in TIMESHEET_PREFIXES:
        if name.startswith(prefix):
            return name[len(prefix):]
    return name


def _check_md_content(md_path: Path) -> list[str]:
    """Spot-check that a per-person timesheet has the expected shape."""
    issues: list[str] = []
    text = md_path.read_text(encoding="utf-8")
    if len(text) < 50:
        issues.append(f"{md_path.name}: file too short ({len(text)} chars)")
        return issues

    # Optional YAML front-matter — if present, must contain iteration/person.
    if text.startswith("---\n"):
        m = re.match(r"^---\n(.*?)\n---", text, re.DOTALL)
        if m:
            block = m.group(1).lower()
            if "iteration" not in block and "person" not in block:
                issues.append(f"{md_path.name}: front-matter missing iteration/person")

    lower = text.lower()
    # Required keywords — the engine writes "Capacity" + "Derived" or
    # the Czech "Kapacita/Odvozeno". Accept either.
    cap_present = any(kw in lower for kw in ("capacity", "kapacita"))
    derived_present = any(kw in lower for kw in ("derived", "odvozeno"))
    if not cap_present:
        issues.append(f"{md_path.name}: missing capacity keyword")
    if not derived_present:
        issues.append(f"{md_path.name}: missing derived keyword")

    return issues


def _load_results_json(path: Path) -> dict | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return {"_error": str(exc)}


def _parse_xlsx(path: Path) -> tuple[tuple[str, ...], list[str]]:
    """Return (sheet names, issues) for a workbook; ``()`` if openpyxl absent."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return (), [f"{path.name}: openpyxl not installed — XLSX content not verified"]

    issues: list[str] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover — corrupt file
        return (), [f"{path.name}: openpyxl failed to open: {exc}"]
    sheets = tuple(wb.sheetnames)
    # Engine writes "Team Summary" + "Item Costs" as of 2.1.x (see Wave B
    # Unit 10 log). If either is missing, flag — but do not treat the
    # presence of additional sheets as a regression.
    required = {"Team Summary", "Item Costs"}
    missing = sorted(required - set(sheets))
    if missing:
        issues.append(f"{path.name}: missing sheets {missing}")
    for sn in sheets:
        ws = wb[sn]
        if ws.max_row < 2:
            issues.append(f"{path.name}: sheet '{sn}' has fewer than 2 rows")
    wb.close()
    return sheets, issues


def _run(cmd: list[str], *, timeout: int = 60) -> tuple[int, str, str]:
    try:
        proc = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout,
            check=False,
        )
    except FileNotFoundError as exc:
        return 127, "", f"command not found: {exc}"
    except subprocess.TimeoutExpired as exc:  # pragma: no cover
        return 124, exc.stdout or "", exc.stderr or "timeout"
    return proc.returncode, proc.stdout, proc.stderr


# -----------------------------------------------------------------------
# Verification steps
# -----------------------------------------------------------------------

def verify_iteration_artifacts(sandbox: Path, result: VerifyResult) -> None:
    reports_dir = sandbox / ".edpa" / "reports"
    if not reports_dir.is_dir():
        result.failures.append(f"reports dir missing: {reports_dir}")
        return

    expected_dirs = {f"iteration-{iter_id}" for iter_id in ITERATIONS}
    actual_dirs = {p.name for p in reports_dir.glob("iteration-PI-*") if p.is_dir()}
    missing_dirs = sorted(expected_dirs - actual_dirs)
    extra_dirs = sorted(actual_dirs - expected_dirs)
    if missing_dirs:
        result.failures.append(f"missing iteration report dirs: {missing_dirs}")
    if extra_dirs:
        # Extras are informational — not a failure (e.g. ad-hoc reruns).
        print(f"INFO: extra iteration dirs present: {extra_dirs}", file=sys.stderr)

    for iter_id in ITERATIONS:
        iter_dir = reports_dir / f"iteration-{iter_id}"
        report = IterationReport(iteration_id=iter_id)

        if not iter_dir.is_dir():
            report.issues.append("directory missing")
            result.iterations.append(report)
            continue

        # 1. Per-person MD timesheets
        md_files = _iter_timesheets(iter_dir, ".md")
        report.md_count = len(md_files)
        seen_people = {_person_id_from_filename(p) for p in md_files}
        missing_people = sorted(set(PEOPLE) - seen_people)
        extra_people = sorted(seen_people - set(PEOPLE))
        if missing_people:
            report.issues.append(f"missing per-person MD for: {missing_people}")
        if extra_people:
            report.issues.append(f"unexpected per-person MD ids: {extra_people}")
        for md in md_files:
            report.issues.extend(_check_md_content(md))

        # 2. Per-person JSON sidecars (optional — engine doesn't write them as
        # of 2.1.2; the JSON contract lives in edpa_results.json). We still
        # count + report so a future addition is captured.
        json_files = _iter_timesheets(iter_dir, ".json")
        report.json_sidecar_count = len(json_files)

        # 3. Team rollup must always exist.
        team_md = iter_dir / "timesheet-team.md"
        if not team_md.is_file():
            # tolerate alternative naming
            team_md = iter_dir / "vykaz-team.md"
        if not team_md.is_file():
            report.issues.append("team rollup timesheet missing")

        # 4. edpa_results.json
        results_path = iter_dir / "edpa_results.json"
        if results_path.is_file():
            report.has_results_json = True
            payload = _load_results_json(results_path)
            if isinstance(payload, dict) and "_error" not in payload:
                report.team_total = payload.get("team_total")
                report.invariants_passed = payload.get("all_invariants_passed")
                if report.invariants_passed is not True:
                    report.issues.append(
                        f"all_invariants_passed != True (got {report.invariants_passed})"
                    )
                # Cross-check: IP iterations should be 0h, others > 0h.
                if iter_id in IP_ITERATIONS and (report.team_total or 0) != 0:
                    report.issues.append(
                        f"IP iteration team_total expected 0, got {report.team_total}"
                    )
                if iter_id not in IP_ITERATIONS and not (report.team_total or 0) > 0:
                    report.issues.append(
                        f"non-IP iteration team_total expected > 0, got {report.team_total}"
                    )
            else:
                report.issues.append(f"edpa_results.json unreadable: {payload}")
        else:
            report.issues.append("edpa_results.json missing")

        # 5. XLSX
        xlsx_files = list(iter_dir.glob("*.xlsx"))
        if xlsx_files:
            report.has_xlsx = True
            sheets, xlsx_issues = _parse_xlsx(xlsx_files[0])
            report.xlsx_sheets = sheets
            report.issues.extend(xlsx_issues)
        else:
            report.issues.append("xlsx export missing")

        result.iterations.append(report)


def verify_pi_summaries(sandbox: Path, result: VerifyResult) -> None:
    reports_dir = sandbox / ".edpa" / "reports"
    for pi in PI_IDS:
        pi_dir = reports_dir / f"pi-{pi}"
        if not pi_dir.is_dir():
            result.failures.append(f"PI dir missing: {pi_dir}")
            continue

        md_files = sorted(p.name for p in pi_dir.glob("*.md"))
        json_files = sorted(p.name for p in pi_dir.glob("*.json"))
        result.pi_summaries[pi] = md_files

        # pi_close.py writes summary.md + pi_results.json
        if "summary.md" not in md_files:
            result.failures.append(f"{pi}: summary.md missing")
        if "pi_results.json" not in json_files:
            result.failures.append(f"{pi}: pi_results.json missing")

        # reports.py --pi writes pi-summary-<PI>.md
        rich_summary = f"pi-summary-{pi}.md"
        if rich_summary not in md_files:
            result.failures.append(f"{pi}: rich summary {rich_summary} missing")
        else:
            text = (pi_dir / rich_summary).read_text(encoding="utf-8")
            # The rich summary must reference all five iterations of the PI.
            iters_for_pi = [iter_id for iter_id in ITERATIONS if iter_id.startswith(pi + ".")]
            missing = [i for i in iters_for_pi if i not in text]
            if missing:
                result.failures.append(
                    f"{pi}: rich summary missing iteration refs: {missing}"
                )


def verify_snapshots(sandbox: Path, result: VerifyResult) -> None:
    snap_dir = sandbox / ".edpa" / "snapshots"
    if not snap_dir.is_dir():
        result.failures.append(f"snapshots dir missing: {snap_dir}")
        return

    snaps = sorted(snap_dir.glob("PI-*.json"))
    result.snapshot_count = len(snaps)
    # 10 iterations + 2 known revisions for PI-2026-1.1 → 12 expected.
    if len(snaps) < len(ITERATIONS):
        result.failures.append(
            f"too few snapshots: {len(snaps)} (expected >= {len(ITERATIONS)})"
        )

    # Each snapshot must be frozen + carry a payload_signature.
    for snap in snaps:
        try:
            payload = json.loads(snap.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            result.failures.append(f"{snap.name}: unreadable ({exc})")
            continue
        if payload.get("frozen") is not True:
            result.failures.append(f"{snap.name}: frozen != True")
        if not payload.get("payload_signature"):
            result.failures.append(f"{snap.name}: payload_signature missing")


def verify_github_state(repo: str, result: VerifyResult) -> None:
    if shutil.which("gh") is None:
        # Non-fatal — the report still has value without GH checks.
        result.failures.append("gh CLI not on PATH — skipping GitHub checks")
        return

    # 1. Merged PRs
    rc, out, err = _run(
        [
            "gh", "pr", "list", "--repo", repo,
            "--state", "merged", "--limit", "200",
            "--json", "number",
        ]
    )
    if rc != 0:
        result.failures.append(f"gh pr list failed (rc={rc}): {err.strip()}")
    else:
        try:
            prs = json.loads(out)
            result.merged_pr_count = len(prs)
            if result.merged_pr_count != EXPECTED_MERGED_PRS:
                result.failures.append(
                    f"merged PR count {result.merged_pr_count} != expected {EXPECTED_MERGED_PRS}"
                )
        except json.JSONDecodeError as exc:
            result.failures.append(f"gh pr list returned invalid JSON: {exc}")

    # 2. CI workflow runs
    rc, out, err = _run(
        [
            "gh", "run", "list", "--repo", repo,
            "--workflow=edpa-contribution-sync.yml",
            "--limit", "100",
            "--json", "conclusion",
        ]
    )
    if rc != 0:
        result.failures.append(f"gh run list failed (rc={rc}): {err.strip()}")
    else:
        try:
            runs = json.loads(out)
            counts: Counter[str | None] = Counter(r.get("conclusion") for r in runs)
            result.ci_run_counts = dict(counts)
            total = sum(counts.values())
            if total < EXPECTED_CI_RUNS_MIN:
                result.failures.append(
                    f"CI workflow runs total {total} < expected min {EXPECTED_CI_RUNS_MIN}"
                )
            # Failure is acknowledged (S-5) — informational only.
        except json.JSONDecodeError as exc:
            result.failures.append(f"gh run list returned invalid JSON: {exc}")

    # 3. Repo archive flag (Wave C: must still be unarchived; Wave D archives).
    rc, out, err = _run(
        ["gh", "repo", "view", repo, "--json", "isArchived", "--jq", ".isArchived"]
    )
    if rc != 0:
        result.failures.append(f"gh repo view failed (rc={rc}): {err.strip()}")
    else:
        flag = out.strip().lower()
        if flag not in ("true", "false"):
            result.failures.append(f"unexpected isArchived value: {flag!r}")
        else:
            result.repo_archived = (flag == "true")
            if result.repo_archived:
                result.failures.append(
                    "repo already archived during Wave C — Wave D should be the one archiving"
                )


# -----------------------------------------------------------------------
# Reporting
# -----------------------------------------------------------------------

def _print_summary(result: VerifyResult, sandbox: Path, repo: str) -> None:
    print("=" * 70)
    print(f"Phase 11 — verify reports + GH state")
    print(f"Sandbox : {sandbox}")
    print(f"GH repo : {repo}")
    print("=" * 70)

    print("\nIteration artifacts:")
    print(
        f"{'Iteration':<14} {'MDs':>3} {'JSON':>4} {'XLSX':>4} "
        f"{'team_h':>7} {'inv_ok':>6} issues"
    )
    for r in result.iterations:
        team_h = "-" if r.team_total is None else f"{r.team_total:.1f}"
        inv_ok = "?" if r.invariants_passed is None else (
            "yes" if r.invariants_passed else "NO"
        )
        xlsx = "yes" if r.has_xlsx else "no"
        n_issues = len(r.issues)
        print(
            f"{r.iteration_id:<14} {r.md_count:>3} {r.json_sidecar_count:>4} "
            f"{xlsx:>4} {team_h:>7} {inv_ok:>6} {n_issues}"
        )
        for issue in r.issues[:5]:
            print(f"   - {issue}")

    print("\nPI summaries:")
    for pi, files in result.pi_summaries.items():
        print(f"  {pi}: {files}")

    print(f"\nFrozen snapshots: {result.snapshot_count}")

    print(f"\nGitHub state:")
    if result.merged_pr_count is not None:
        print(f"  merged PRs           : {result.merged_pr_count}")
    if result.ci_run_counts:
        print(f"  CI workflow runs     : {result.ci_run_counts}")
    if result.repo_archived is not None:
        print(f"  repo isArchived      : {result.repo_archived}")

    print(f"\nResult: {'PASS' if result.passed else 'FAIL'} ({len(result.failures)} failures)")
    if result.failures:
        print("Failures:")
        for f in result.failures:
            print(f"  - {f}")


# -----------------------------------------------------------------------
# Entry point
# -----------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    sandbox = _resolve_sandbox()
    repo = _resolve_repo()

    if not sandbox.is_dir():
        print(f"ERROR: sandbox dir not found: {sandbox}", file=sys.stderr)
        return 2

    result = VerifyResult()
    verify_iteration_artifacts(sandbox, result)
    # Per-iteration issues are warnings; promote to failures only when an
    # iteration has *any* issue (so the run log is precise).
    for r in result.iterations:
        if r.issues:
            result.failures.extend(f"{r.iteration_id}: {i}" for i in r.issues)

    verify_pi_summaries(sandbox, result)
    verify_snapshots(sandbox, result)
    verify_github_state(repo, result)

    _print_summary(result, sandbox, repo)
    return 0 if result.passed else 1


if __name__ == "__main__":
    sys.exit(main())
