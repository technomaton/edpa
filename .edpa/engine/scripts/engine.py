#!/usr/bin/env python3
"""
EDPA Engine — Evidence-Driven Proportional Allocation

Standalone Python implementation of the EDPA calculation engine.
Computes derived hours from delivery evidence stored in .edpa/backlog/.

Usage:
    python3 .edpa/engine/scripts/engine.py --edpa-root .edpa --iteration PI-2026-1.3
    python3 .edpa/engine/scripts/engine.py --demo  # Run with built-in sample data

    # Legacy mode (requires external item gathering):
    python3 .edpa/engine/scripts/engine.py --capacity cap.yaml --heuristics h.yaml --iteration PI-2026-1.3
"""

try:  # best-effort UTF-8 stdio on legacy Windows consoles (cp1250)
    import _console  # noqa: F401
except ImportError:
    pass
import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import yaml
except ImportError:
    print("ERROR: PyYAML required. Install with: pip install pyyaml")
    sys.exit(1)


def get_version():
    """Read version from plugin.json or the vendored VERSION file."""
    # plugin.json candidates (EDPA repo: plugin/.claude-plugin/plugin.json)
    for candidate in [
        Path(__file__).parent.parent.parent / ".claude-plugin" / "plugin.json",
        Path(__file__).parent.parent.parent.parent / ".claude" / ".claude-plugin" / "plugin.json",
    ]:
        try:
            if candidate.exists():
                with open(candidate, encoding="utf-8") as f:
                    return json.load(f).get("version", "unknown")
        except (json.JSONDecodeError, OSError):
            continue
    # Installed project: .edpa/engine/VERSION written by install.sh
    version_file = Path(__file__).parent.parent / "VERSION"
    try:
        if version_file.exists():
            return version_file.read_text(encoding="utf-8").strip()
    except OSError:
        pass
    return "unknown"


VERSION = get_version()

# Evidence roles — kept as a constant for backward compatibility with
# pre-v1.11 callers (validate_syntax.py, migrate_contributors.py). In
# v1.11 the engine no longer uses role-based CW computation; cw values
# are pre-computed by detect_contributors.py via per-item normalization
# and consumed directly. This constant remains for tooling that still
# emits role labels (e.g., display layer in reports.py).
EVIDENCE_ROLES = {"owner", "key", "reviewer", "consulted"}


sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from _yaml_io import load_yaml  # noqa: E402  (shared .md/.yaml loader, S-242)
finally:
    sys.path.pop(0)


def gh_json(cmd):
    """Run gh CLI command and parse JSON output."""
    try:
        result = subprocess.run(
            ["gh"] + cmd.split() + ["--json", "number,title,assignees,labels,body"],
            capture_output=True, text=True, timeout=30, encoding="utf-8"
        )
        if result.returncode == 0:
            return json.loads(result.stdout)
    except (subprocess.TimeoutExpired, FileNotFoundError, json.JSONDecodeError):
        pass
    return []


def extract_item_refs(text):
    """Extract work item references (S-123, F-45, E-7) from text."""
    if not text:
        return []
    return re.findall(r'[SFEITD]-\d+', text)


def extract_contributors(item):
    """Extract pre-computed (person_id, cw, signals) tuples from an item.

    v1.11: detect_contributors.py is the single source for CW
    computation. Each contributors[] entry already has:
      - person      (canonical id from people.yaml)
      - cw          (per-item-normalized share, [0,1])
      - contribution_score (raw sum of signal weights, ≥ 0)
      - signals[]   (audit trail with type/ref/weight/...)

    Engine reads cw verbatim — no role mapping, no signal scoring,
    no /contribute regex. This function exists only to extract the
    tuple format the rest of run_edpa expects, with safe handling
    of legacy fixtures (older `as:` field is ignored, cw still read).

    Returns: list of {"person": str, "cw": float, "signals": list}
    """
    out = []
    for c in (item.get("contributors") or []):
        if not isinstance(c, dict):
            continue
        person = c.get("person")
        cw = c.get("cw")
        if not person or cw is None:
            continue
        try:
            cw_val = float(cw)
        except (TypeError, ValueError):
            continue
        if not 0 <= cw_val <= 1:
            continue
        out.append({
            "person": person,
            "cw": cw_val,
            "signals": c.get("signals", []),
        })
    return out


def _enrich_items_with_yaml_edit_signals(items, yaml_edit_signals, people):
    """v1.17: merge yaml_edit signals into items' contributors[] in-memory.

    For each (item, person) pair, add yaml_edit weights to the existing
    contribution_score (or seed a new entry if the person wasn't
    previously credited), then re-normalize cw across all persons for
    that item. The mutation is in-memory only — YAML files on disk are
    not touched. The frozen snapshot captures the augmented contributors
    so the audit trail is complete.

    Item id resolution: gate_event ids look like `F-7@Funnel→Analyzing`;
    we strip the `@...` suffix so all gate_events of the same parent
    share the parent's enriched contributors[].
    """
    from collections import defaultdict

    # Build login → person_id resolver. People may register one of:
    #   github (login),  email,  or just id. yaml_edit_signals.py emits
    #   `login` populated with whichever resolved at the commit-author
    #   level, so we accept all three.
    resolver = {}
    for p in people:
        pid = p.get("id")
        if not pid:
            continue
        for key in ("github", "email", "id"):
            v = p.get(key)
            if v:
                resolver[str(v).lower()] = pid

    for item in items:
        raw_id = item.get("id", "")
        item_id = raw_id.split("@", 1)[0]  # strip gate-event suffix
        sigs = yaml_edit_signals.get(item_id) or []
        if not sigs:
            continue

        # Aggregate yaml_edit weight + signals per person.
        weight_per_person = defaultdict(float)
        signals_per_person = defaultdict(list)
        for s in sigs:
            login = (s.get("login") or "").lower()
            person_id = resolver.get(login)
            if not person_id:
                # Unknown commit author — skip silently; auditor sees the
                # commit ref in the iteration's signal log either way.
                continue
            weight_per_person[person_id] += float(s.get("weight", 0))
            signals_per_person[person_id].append(s)

        if not weight_per_person:
            continue

        # Existing contributors (from detect_contributors or seed). Their
        # contribution_score is the canonical raw weight. When missing
        # (legacy YAMLs), fall back to cw — this means cw is treated as
        # if it were already a score, which preserves relative shares.
        existing = item.get("contributors") or []
        contrib_score = {}
        signal_pool: dict[str, list] = {}
        for c in existing:
            pid = c.get("person")
            if not pid:
                continue
            base = c.get("contribution_score")
            if base is None:
                base = float(c.get("cw", 0))
            contrib_score[pid] = float(base)
            signal_pool[pid] = list(c.get("signals", []) or [])

        # Stack yaml_edit weights on top.
        for pid, yaml_w in weight_per_person.items():
            contrib_score[pid] = contrib_score.get(pid, 0) + yaml_w
            signal_pool.setdefault(pid, []).extend(signals_per_person[pid])

        total = sum(s for s in contrib_score.values() if s > 0)
        if total <= 0:
            continue

        item["contributors"] = [
            {
                "person": pid,
                "cw": round(score / total, 4),
                "contribution_score": round(score, 2),
                "signals": signal_pool.get(pid, []),
            }
            for pid, score in contrib_score.items()
            if score > 0
        ]


def _load_iteration_people_overrides(edpa_root, iteration_id):
    """Read iteration-level `people:` overrides from
    .edpa/iterations/<id>.yaml. Reuses the same schema as
    .edpa/config/people.yaml; only fields explicitly set on the
    iteration entry override the baseline (matching by `id`).

    Returns a dict keyed by person id of the override fields, or an
    empty dict when the iteration file has no `people:` section (or
    the file doesn't exist). Optional `note` is preserved for the
    audit trail (snapshot + reports) but doesn't affect the math.

    Schema (v1.9.0+):
        iteration:
          id: PI-2026-1.3
          ...
        people:
          - id: bob-dev
            capacity_per_iteration: 44
            note: "IP weekend deploy push (Jun 13-14)"
          - id: alice-arch
            capacity_per_iteration: 10
            note: "vacation Jun 9-11"
    """
    if not edpa_root or not iteration_id:
        return {}
    iter_file = Path(edpa_root) / "iterations" / f"{iteration_id}.yaml"
    if not iter_file.is_file():
        return {}
    try:
        data = yaml.safe_load(iter_file.read_text(encoding="utf-8")) or {}
    except (yaml.YAMLError, OSError):
        return {}
    entries = data.get("people") or []
    out = {}
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        pid = entry.get("id")
        if not pid:
            continue
        # Store the full entry so future fields (availability, etc.)
        # can be overridden without changing this loader. Engine reads
        # the keys it knows about; everything else is just metadata.
        out[pid] = dict(entry)
    return out


# Set of people.yaml fields that the engine knows how to override per
# iteration. Anything else in iteration.people[<entry>] is preserved as
# metadata in the snapshot but doesn't affect the calculation.
ITERATION_OVERRIDABLE_FIELDS = {"capacity_per_iteration", "capacity"}


def _resolve_capacity(person, override):
    """Apply iteration-level override on top of the baseline capacity
    declared in people.yaml.

    Returns (effective, baseline, override_meta) where:
      - effective is the capacity to feed into the proportional allocation
      - baseline is people[].capacity_per_iteration from people.yaml
      - override_meta is None when no override was applied, otherwise a
        dict {capacity, note?} suitable for snapshot persistence

    Raises ValueError when the resolved capacity would be negative —
    the engine treats this as a configuration error rather than
    silently clamping to 0.
    """
    cpi = person.get("capacity_per_iteration")
    baseline = cpi if cpi is not None else person.get("capacity", 0)
    if not override:
        return baseline, baseline, None
    eff_source = None
    eff = baseline
    if override.get("capacity_per_iteration") is not None:
        eff = float(override["capacity_per_iteration"])
        eff_source = "capacity_per_iteration"
    elif override.get("capacity") is not None:
        eff = float(override["capacity"])
        eff_source = "capacity"
    if eff_source is None:
        # Override entry exists but doesn't touch capacity (e.g., only
        # `note:`). Effective stays at baseline; we still record the
        # note so reports can surface it.
        if override.get("note"):
            return baseline, baseline, {
                "capacity": baseline,
                "note": override.get("note", ""),
            }
        return baseline, baseline, None
    if eff < 0:
        raise ValueError(
            f"iteration people override for {person.get('id', '?')!r} "
            f"produces negative capacity ({eff}h); check the value"
        )
    return eff, baseline, {
        "capacity": eff,
        "note": override.get("note", ""),
    }


def run_edpa(capacity_config, heuristics, items, *,
             edpa_root=None, iteration_id=None):
    """
    Run the core EDPA calculation.

    v1.14: single calculation path. The pre-v1.14 `mode` parameter
    (simple|full|gates) was removed — gates is the only mode and is
    a strict superset of simple (Done items + transition events when
    git history records them; degenerates to Done-only if no
    transitions exist). For backward compatibility v1.13 callers may
    still pass mode= as a kwarg; it's accepted and ignored.

    Returns: list of person results with derived hours.

    edpa_root + iteration_id are optional but required together when
    capacity_overrides should be applied — engine reads them from
    .edpa/iterations/<id>.yaml. When either is None, behavior matches
    pre-v1.9 (every person uses people.yaml baseline only).
    """
    people = capacity_config.get("people", [])
    overrides_map = _load_iteration_people_overrides(edpa_root, iteration_id)

    # v1.11: pre-build (person, item) → (cw, signals) lookup from
    # contributors[] blocks already populated by detect_contributors.py.
    # No evidence detection, no CW computation here — engine is a thin
    # consumer of pre-aggregated values.
    contributors_by_item: dict[str, list[dict]] = {}
    item_contribution_total: dict[str, float] = {}
    for item in items:
        contribs = extract_contributors(item)
        contributors_by_item[item["id"]] = contribs
        # Sanity: cw shares should sum to ~1.0 per item. We log but
        # don't fail — re-running detect_contributors fixes drift.
        s = sum(c["cw"] for c in contribs)
        item_contribution_total[item["id"]] = s
        if contribs and abs(s - 1.0) > 0.01:
            print(f"WARN: {item['id']}: Σ contributors[].cw = {s:.4f} "
                  f"(expected 1.0). Re-run detect_contributors to recompute.",
                  file=sys.stderr)

    results = []

    for person in people:
        pid = person["id"]
        capacity, baseline, override_meta = _resolve_capacity(
            person, overrides_map.get(pid))
        person_items = []

        for item in items:
            item_id = item["id"]
            js = item.get("job_size", 0)
            if js <= 0:
                continue

            # Find this person's pre-computed cw on this item.
            person_cw = None
            person_signals = []
            for c in contributors_by_item.get(item_id, []):
                if c["person"] == pid:
                    person_cw = c["cw"]
                    person_signals = c["signals"]
                    break
            if person_cw is None or person_cw <= 0:
                continue  # no contribution to this item

            # v1.11: cw is already per-item-normalized. score = JS × cw
            # gives the person's effective work units on this item;
            # `rs` (relevance signal) was a v1.7-era refinement that
            # is now redundant — multi-signal evidence is already
            # baked into cw via the additive aggregation in detect.
            score = js * person_cw
            rs = 1.0

            person_items.append({
                "id": item_id,
                "level": item.get("level", "Story"),
                "js": js,
                "cw": round(person_cw, 4),
                "rs": round(rs, 4),
                "score": round(score, 4),
                "evidence": [s.get("type", "?") for s in person_signals],
            })

        # Calculate derived hours
        sum_scores = sum(pi["score"] for pi in person_items)

        for pi in person_items:
            if sum_scores > 0:
                ratio = pi["score"] / sum_scores
                hours = ratio * capacity
            else:
                ratio = 0.0
                hours = 0.0
            pi["ratio"] = round(ratio, 6)
            pi["hours"] = round(hours, 2)

        # Normalize: adjust last item so sum exactly equals capacity
        if person_items and sum_scores > 0:
            rounded_sum = sum(pi["hours"] for pi in person_items)
            diff = round(capacity - rounded_sum, 2)
            if diff != 0:
                person_items[-1]["hours"] = round(person_items[-1]["hours"] + diff, 2)

        total_derived = sum(pi["hours"] for pi in person_items)

        # Validate invariants
        invariant_ok = True
        if person_items:
            if abs(total_derived - capacity) > 0.1:
                invariant_ok = False
            ratio_sum = sum(pi["ratio"] for pi in person_items)
            if abs(ratio_sum - 1.0) > 0.001:
                invariant_ok = False
            if any(pi["hours"] < 0 for pi in person_items):
                invariant_ok = False

        result_entry = {
            "id": pid,
            "name": person.get("name", pid),
            "role": person.get("role", ""),
            "capacity": capacity,
            "total_derived": round(total_derived, 2),
            "items": person_items,
            "invariant_ok": invariant_ok,
        }
        # Surface override metadata only when an override was actually
        # applied — keeps pre-1.9 snapshots byte-identical for runs that
        # don't touch capacity_overrides.
        if override_meta is not None:
            result_entry["capacity_baseline"] = baseline
            result_entry["capacity_override"] = override_meta
        results.append(result_entry)

    return results


def load_heuristics(edpa_root):
    """Load CW heuristics from .edpa/config/, with fallback chain.

    Tries (in order):
      1. .edpa/config/heuristics.yaml         (legacy v1.x)
      2. .edpa/config/cw_heuristics.yaml      (V2.1 seeded by project_setup)
      3. .edpa/engine/templates/cw_heuristics.yaml.tmpl  (V2 vendored)
      4. .claude/edpa/templates/cw_heuristics.yaml.tmpl  (legacy V1 location)
      5. Hardcoded minimal default (no gate_weights — gate events skipped)

    Without one of 1-4 hitting, the engine loses gate_weights and
    yaml_edit_weights — only signal_weights from contributors[] are
    used. Krok C7 added (2) seeding so the typical install reaches the
    documented defaults from path 2.
    """
    edpa_root = Path(edpa_root)
    for name in ("heuristics.yaml", "cw_heuristics.yaml"):
        path = edpa_root / "config" / name
        if path.exists():
            return load_yaml(path)
    # Fallbacks: V2 vendored layout first, then legacy V1 path.
    candidates = [
        edpa_root / "engine" / "templates" / "cw_heuristics.yaml.tmpl",
        edpa_root.parent / ".claude" / "edpa" / "templates" / "cw_heuristics.yaml.tmpl",
    ]
    for template in candidates:
        if template.exists():
            return load_yaml(template)
    return {"evidence_threshold": 1.0, "role_weights": {"owner": 1.0, "key": 0.6, "reviewer": 0.25, "consulted": 0.15}}


def load_backlog_items(edpa_root, iteration_id=None):
    """Read .edpa/backlog/ YAML files and convert to engine item format.

    Each backlog YAML has: id, type, title, js, status, assignee, contributors
    Engine expects: id, level, job_size, assignees, contributors

    Args:
        edpa_root: Path to .edpa/ directory
        iteration_id: If given, only include items matching this iteration. If None, include all Done items.

    Returns:
        List of item dicts in engine format, plus a dict of manual CW overrides.
    """
    edpa_root = Path(edpa_root)
    backlog_dir = edpa_root / "backlog"
    if not backlog_dir.exists():
        return [], {}

    items = []
    manual_cw_overrides = {}  # {(person_id, item_id): cw_value}
    schema_warnings = []      # collected per-item schema problems
    contributors_seen_total = 0
    evidence_pairs_total = 0

    type_dirs = {
        "stories": "Story",
        "features": "Feature",
        "epics": "Epic",
        "initiatives": "Initiative",
        "defects": "Defect",
    }

    for dir_name, level in type_dirs.items():
        type_dir = backlog_dir / dir_name
        if not type_dir.exists():
            continue

        for md_file in sorted(type_dir.glob("*.md")):
            data = load_yaml(md_file)
            if data is None:
                continue

            if not data or not isinstance(data, dict):
                continue

            item_id = data.get("id", md_file.stem)
            status = data.get("status", "")

            # Filter: only Done items
            if status.lower() not in ("done", "closed", "accepted"):
                continue

            # Filter by iteration — SAFe hierarchy-aware:
            #   Story / Defect / Task → exact iteration match (e.g., PI-2026-1.1)
            #   Feature → PI match (e.g., PI-2026-1 matches PI-2026-1.x)
            #   Epic/Initiative → always included if Done (cross-PI)
            item_type = data.get("type", level)
            item_iter = data.get("iteration", "")

            if iteration_id:
                if item_type in ("Story", "Defect", "Task"):
                    if item_iter != iteration_id:
                        continue
                elif item_type == "Feature":
                    pi_prefix = iteration_id.rsplit(".", 1)[0]
                    if item_iter != pi_prefix and item_iter != iteration_id:
                        continue
                # Epic + Initiative: always included if Done

            js = data.get("js") or data.get("job_size", 0)
            if not js or js <= 0:
                continue

            # v1.11: contributors[] is the single source of truth.
            # Each entry has person + cw (per-item share) + signals[].
            # Engine consumes cw verbatim; no role mapping, no /contribute
            # body synthesis.
            raw_contribs = data.get("contributors", []) or []
            contributors = []
            for idx, contrib in enumerate(raw_contribs):
                if not isinstance(contrib, dict):
                    schema_warnings.append(
                        f"{item_id}: contributors[{idx}] is not a mapping "
                        f"(got {type(contrib).__name__})"
                    )
                    continue
                contributors_seen_total += 1
                person = contrib.get("person", "")
                if not person:
                    schema_warnings.append(
                        f"{item_id}: contributors[{idx}] missing 'person'"
                    )
                    continue

                # Reject legacy keys with a migration breadcrumb. v1.11
                # is a hard schema cut — old fields cannot be silently
                # interpreted.
                if "role" in contrib:
                    schema_warnings.append(
                        f"{item_id}: contributors[{idx}] uses legacy "
                        f"'role' — rewritten in v1.11 (signals[] now "
                        f"carries audit trail; cw carries the share). "
                        f"Run detect_contributors.py to regenerate."
                    )
                    continue
                if "weight" in contrib:
                    schema_warnings.append(
                        f"{item_id}: contributors[{idx}] uses legacy "
                        f"'weight' — replaced by 'cw' (since v1.7) and "
                        f"'contribution_score' (since v1.11). "
                        f"Run detect_contributors.py to regenerate."
                    )
                    continue
                # Tolerate the v1.10 `as:` field for one transition: ignored
                # but doesn't cause skip. Validator (validate_syntax.py)
                # still rejects it on commit-time hooks.

                cw = contrib.get("cw")
                if cw is None:
                    schema_warnings.append(
                        f"{item_id}: contributors[{idx}] missing 'cw' "
                        f"(per-item share). Run detect_contributors.py."
                    )
                    continue
                try:
                    cw_val = float(cw)
                except (TypeError, ValueError):
                    schema_warnings.append(
                        f"{item_id}: contributors[{idx}] cw must be numeric"
                    )
                    continue

                contributors.append({
                    "person": person,
                    "cw": cw_val,
                    "contribution_score": contrib.get("contribution_score", 0),
                    "signals": contrib.get("signals", []),
                })
                evidence_pairs_total += 1

            # Top-level assignees are tracked for the snapshot only —
            # they do not feed evidence detection or scoring. (The
            # assignee signal was removed; CW comes from commit / review
            # / comment evidence in contributors[].signals.)
            assignees = []
            top_assignees = data.get("assignees") or []
            if isinstance(top_assignees, list):
                for a in top_assignees:
                    if isinstance(a, dict) and a.get("login"):
                        assignees.append({"login": a["login"]})
                    elif isinstance(a, str):
                        assignees.append({"login": a})
            assignee = data.get("assignee") or data.get("owner")
            if assignee and not any(a.get("login") == assignee for a in assignees):
                assignees.append({"login": assignee})

            items.append({
                "id": item_id,
                "level": data.get("type", level),
                "job_size": js,
                "assignees": assignees,
                "contributors": contributors,
            })

    if schema_warnings:
        print("", file=sys.stderr)
        print("WARN: backlog schema issues detected:", file=sys.stderr)
        for w in schema_warnings:
            print(f"  - {w}", file=sys.stderr)
        print("", file=sys.stderr)

    if contributors_seen_total > 0 and evidence_pairs_total == 0:
        print(
            "WARN: 0 evidence pairs derived from "
            f"{contributors_seen_total} contributor entries. "
            "Engine will allocate 0h. Check contributors[].cw is set "
            "for every entry — v1.11 schema requires per-item-normalized "
            "cw shares produced by detect_contributors.py.",
            file=sys.stderr,
        )

    return items, manual_cw_overrides


GATE_TYPE_DIRS = {
    "Feature": "features",
    "Epic": "epics",
    "Initiative": "initiatives",
}


def _passthrough_contributors(item_data):
    """v1.11: pass parent's contributors[] through to gate event verbatim.

    The pre-v1.11 mapping (as: → top-level evidence fields) is gone —
    engine reads cw directly from contributors[]. Gate events synthesised
    from parent transitions inherit the parent's contributors with their
    pre-computed shares, which is the correct semantic: whoever drove
    the parent across its lifecycle gets credited at each gate.
    """
    contribs = []
    for c in (item_data.get("contributors") or []):
        if not isinstance(c, dict):
            continue
        person = c.get("person")
        cw = c.get("cw")
        if not person or cw is None:
            continue
        contribs.append({
            "person": person,
            "cw": cw,
            "contribution_score": c.get("contribution_score", 0),
            "signals": c.get("signals", []),
        })
    return contribs


def _build_person_resolver(people):
    """Build login/email/id → person_id resolver from people.yaml entries."""
    resolver = {}
    for p in people or []:
        pid = p.get("id")
        if not pid:
            continue
        for key in ("github", "email", "id"):
            v = p.get(key)
            if v:
                resolver[str(v).lower()] = pid
    return resolver


def load_story_activity_events(edpa_root, iteration_id, heuristics,
                                yaml_edit_signals):
    """V2.1 C7.5 — emit synthetic items for in-flight Stories with activity.

    Stories normally credit only at status=Done (full Story.js × cw)
    via load_backlog_items. That loses refinement work on Stories that
    spill across iterations: yaml_edit signals exist but the Story
    isn't in items[] yet, so enrichment skips it.

    This function emits one synthetic item per (in-flight Story with
    yaml_edit_signals in the iter window). The synthetic item has:
      id        = f"{story_id}@activity"
      level     = "Story"
      job_size  = story.js * credit_factor    (configurable; default 0.40)
      contributors = []  (filled by _enrich_items_with_yaml_edit_signals)

    The credit_factor is reserved capacity for "in-progress activity"
    — a refinement-heavy iteration crediting 40 % of Story.js per
    contributor matches the empirical observation that prep + grooming
    typically consume 30–50 % of a Story's overall delivery effort.
    Tune via ``story_activity.credit_factor`` in cw_heuristics.yaml.

    Skip rules:
      - status=Done             → load_backlog_items already credits
      - no yaml_edit signals     → no activity to credit
      - story.js <= 0            → engine can't score zero-js items
      - factor <= 0              → C7.5 disabled by config

    Returns (events, audit). events go into items[]; audit logged into
    edpa_results.json for inspection.
    """
    if not iteration_id or not yaml_edit_signals:
        return [], []
    factor = float(((heuristics or {}).get("story_activity") or {}).get(
        "credit_factor", 0.40))
    if factor <= 0:
        return [], []

    stories_dir = Path(edpa_root) / "backlog" / "stories"
    if not stories_dir.is_dir():
        return [], []

    events = []
    audit = []
    for story_path in sorted(stories_dir.glob("*.md")):
        data = load_yaml(story_path) or {}
        story_id = data.get("id")
        if not story_id:
            continue
        if (data.get("status") or "").lower() in ("done", "closed", "accepted"):
            continue
        sigs = yaml_edit_signals.get(story_id) or []
        if not sigs:
            continue
        js = data.get("js") or data.get("job_size") or 0
        if js <= 0:
            continue

        events.append({
            "id": f"{story_id}@activity",
            "level": "Story",
            "job_size": round(js * factor, 4),
            "title": data.get("title", ""),
            "contributors": [],  # populated by yaml_edit enrichment
        })
        audit.append({
            "item_id": story_id,
            "type": "story_activity",
            "credit_factor": factor,
            "story_js": js,
            "effective_js": round(js * factor, 4),
            "n_yaml_edit_signals": len(sigs),
        })
    return events, audit


def load_gate_events(edpa_root, iteration_id, heuristics, people=None):
    """Convert status transitions into scoring 'events' for mode=gates.

    For Feature/Epic/Initiative parents, every status transition that occurred
    within iteration_id's date window becomes an item-shaped event with
    job_size = parent.js * gate_weights[type][transition]. Each event reuses
    its parent's contributor list as evidence, so run_edpa() scores it with
    the same math as a Done item.

    v1.17.1 fix (Finding #1): when the parent has no contributors[] (typical
    for IP-iter strategic items seeded with title+js but no team yet), fall
    back to crediting the transition's commit author (`changed_by`) at cw=1.0.
    Without this fallback, IP iterations with real strategic work derive 0h
    because gate events inherit empty contributor lists.

    Stories are NOT emitted here — they continue to flow through
    load_backlog_items() with the Done filter.
    """
    edpa_root = Path(edpa_root)
    if not iteration_id:
        return [], []
    resolver = _build_person_resolver(people or [])

    iter_file = edpa_root / "iterations" / f"{iteration_id}.yaml"
    if not iter_file.is_file():
        return [], []

    sys.path.insert(0, str(Path(__file__).parent))
    try:
        from transitions import parse_iteration_dates, detect_transitions
    finally:
        sys.path.pop(0)

    try:
        start, end = parse_iteration_dates(iter_file)
    except (ValueError, KeyError) as e:
        print(f"WARN: cannot parse iteration dates: {e}", file=sys.stderr)
        return [], []

    transitions = detect_transitions(edpa_root, since=start, until=end)
    gate_weights = (heuristics or {}).get("gate_weights", {}) or {}

    events = []
    audit = []
    for t in transitions:
        item_type = t["item_type"]
        # Stories are surfaced by transitions.py for audit/debug visibility,
        # but engine gates mode credits stories only at status=Done (handled
        # in load_backlog_items). Skip story-level transitions here so we
        # don't double-count.
        if item_type == "Story":
            continue
        weights = gate_weights.get(item_type, {}) or {}
        gate_key = f"{t['from_status']}→{t['to_status']}"
        weight = weights.get(gate_key)
        if weight is None and weights:
            weight = round(1.0 / len(weights), 4)
            print(
                f"WARN: no gate_weight for {item_type} '{gate_key}', "
                f"using equal-split fallback {weight}",
                file=sys.stderr,
            )
        if not weight or weight <= 0:
            continue

        sub = GATE_TYPE_DIRS.get(item_type)
        if not sub:
            continue
        parent_file = edpa_root / "backlog" / sub / f"{t['item_id']}.md"
        if not parent_file.is_file():
            continue
        parent = load_yaml(parent_file) or {}
        parent_js = parent.get("js") or parent.get("job_size") or 0
        if parent_js <= 0:
            continue

        effective_js = round(parent_js * weight, 6)
        synth_id = f"{t['item_id']}@{t['from_status'] or 'init'}->{t['to_status']}"

        contribs = _passthrough_contributors(parent)
        if not contribs:
            # v1.17.1 fallback: parent has no contributors[] yet → credit the
            # transition's git author. Resolves email/login via people.yaml.
            changed_by = (t.get("changed_by") or "").lower()
            resolved = resolver.get(changed_by)
            # Some signals carry just the local-part of email (e.g. "jurby"
            # from "jurby@noreply.github.com"). Try a relaxed match too.
            if not resolved and "@" in changed_by:
                resolved = resolver.get(changed_by.split("@", 1)[0])
            if resolved:
                contribs = [{
                    "person": resolved,
                    "cw": 1.0,
                    "contribution_score": float(weight),
                    "signals": [{
                        "type": "gate_transition_author",
                        "ref": t.get("commit_hash"),
                        "weight": float(weight),
                        "transition": gate_key,
                        "parent_id": t["item_id"],
                        "detected_at": t.get("changed_at"),
                    }],
                }]

        events.append({
            "id": synth_id,
            "level": item_type,
            "job_size": effective_js,
            "contributors": contribs,
        })
        audit.append({
            "synth_id": synth_id,
            "parent_id": t["item_id"],
            "parent_type": item_type,
            "transition": gate_key,
            "weight": weight,
            "effective_js": effective_js,
            "changed_at": t["changed_at"],
            "changed_by": t["changed_by"],
            "commit_hash": t["commit_hash"],
        })

    return events, audit


def generate_demo_data():
    """Generate sample data for demonstration (multi-contract).

    Alice is split into two contracts:
      - alice-arch  (Arch, 40h) — scoped to Stories (S-*), evidence_default=true
      - alice-pm    (PM,  20h) — scoped to Epics/Features (E-*, F-*)
    Total team capacity: 40 + 20 + 80 + 60 = 200h.
    """
    capacity = {
        "teams": [
            {"id": "Alpha", "planning_factor": 0.8},
        ],
        "people": [
            {"id": "alice-arch", "name": "Alice (Arch)", "role": "Arch", "team": "Alpha",
             "fte": 0.5, "capacity_per_iteration": 40, "email": "alice@example.com",
             "evidence_scope": ["S-*"], "evidence_default": True},
            {"id": "alice-pm", "name": "Alice (PM)", "role": "PM", "team": "Alpha",
             "fte": 0.25, "capacity_per_iteration": 20, "email": "alice@example.com",
             "evidence_scope": ["E-*", "F-*"]},
            {"id": "bob", "name": "Bob (Dev)", "role": "Dev", "team": "Alpha",
             "fte": 1.0, "capacity_per_iteration": 80, "email": "bob@example.com"},
            {"id": "carol", "name": "Carol (Dev)", "role": "Dev", "team": "Alpha",
             "fte": 0.75, "capacity_per_iteration": 60, "email": "carol@example.com"},
        ]
    }

    heuristics = {
        "version": "2.2",
        # Flat signal weights (no role_overrides or role_weights; those
        # were dropped). The GitHub-issue ``assignee`` and ``pr_author``
        # signals were removed — local-first attribution rests on
        # commit_author, pr_reviewer and issue_comment.
        "signals": {"commit_author": 2.78, "pr_reviewer": 2.25,
                    "issue_comment": 1.14},
    }

    # Demo items: contributors[] with pre-computed cw shares (per-item,
    # Σ = 1.0). Engine reads these directly. The cw values below are
    # illustrative — they reflect what detect_contributors.py would
    # produce from the commit / review / comment signals listed under
    # signals[] for each contributor (owners dominate via more commits;
    # a PM with no commits is credited through issue_comments).
    items = [
        # S-101: bob owns via commits, carol co-commits, alice-arch reviews
        {"id": "S-101", "level": "Story", "job_size": 5,
         "assignees": [{"login": "bob"}],
         "contributors": [
             {"person": "bob", "cw": 0.624, "contribution_score": 8.34,
              "signals": [{"type": "commit_author", "ref": "S-101/commit/aa1", "weight": 2.78},
                          {"type": "commit_author", "ref": "S-101/commit/aa2", "weight": 2.78},
                          {"type": "commit_author", "ref": "S-101/commit/aa3", "weight": 2.78}]},
             {"person": "carol", "cw": 0.208, "contribution_score": 2.78,
              "signals": [{"type": "commit_author", "ref": "S-101/commit/aa4", "weight": 2.78}]},
             {"person": "alice-arch", "cw": 0.168, "contribution_score": 2.25,
              "signals": [{"type": "pr_reviewer", "ref": "pr#1/review/r1", "weight": 2.25}]},
         ]},
        # S-102: carol owns via commits; bob co-commits + reviews; alice-arch consulted (manual + comment)
        {"id": "S-102", "level": "Story", "job_size": 8,
         "assignees": [{"login": "carol"}],
         "contributors": [
             {"person": "carol", "cw": 0.538, "contribution_score": 8.34,
              "signals": [{"type": "commit_author", "ref": "S-102/commit/bb1", "weight": 2.78},
                          {"type": "commit_author", "ref": "S-102/commit/bb2", "weight": 2.78},
                          {"type": "commit_author", "ref": "S-102/commit/bb3", "weight": 2.78}]},
             {"person": "bob", "cw": 0.324, "contribution_score": 5.03,
              "signals": [{"type": "pr_reviewer", "ref": "pr#2/review/r2", "weight": 2.25},
                          {"type": "commit_author", "ref": "S-102/commit/bb4", "weight": 2.78}]},
             {"person": "alice-arch", "cw": 0.138, "contribution_score": 2.14,
              "signals": [{"type": "manual:pr_body", "ref": "pr#2/body",
                           "excerpt": "/contribute @alice-arch weight:0.6", "weight": 0.60},
                          {"type": "issue_comment", "ref": "issue#102/comment/c1", "weight": 1.14},
                          {"type": "manual:pr_body", "ref": "pr#2/body", "weight": 0.4}]},
         ]},
        # S-103: bob solo
        {"id": "S-103", "level": "Story", "job_size": 3,
         "assignees": [{"login": "bob"}],
         "contributors": [
             {"person": "bob", "cw": 0.787, "contribution_score": 8.34,
              "signals": [{"type": "commit_author", "ref": "S-103/commit/cc1", "weight": 2.78},
                          {"type": "commit_author", "ref": "S-103/commit/cc2", "weight": 2.78},
                          {"type": "commit_author", "ref": "S-103/commit/cc3", "weight": 2.78}]},
             {"person": "alice-arch", "cw": 0.213, "contribution_score": 2.25,
              "signals": [{"type": "pr_reviewer", "ref": "pr#3/review/r3", "weight": 2.25}]},
         ]},
        # F-10: alice-pm leads the feature via issue comments (no commits), devs comment too
        {"id": "F-10", "level": "Feature", "job_size": 13,
         "assignees": [{"login": "alice-pm"}],
         "contributors": [
             {"person": "alice-pm", "cw": 0.60, "contribution_score": 3.42,
              "signals": [{"type": "issue_comment", "ref": "issue#10/comment/c1", "weight": 1.14},
                          {"type": "issue_comment", "ref": "issue#10/comment/c2", "weight": 1.14},
                          {"type": "issue_comment", "ref": "issue#10/comment/c3", "weight": 1.14}]},
             {"person": "bob", "cw": 0.20, "contribution_score": 1.14,
              "signals": [{"type": "issue_comment", "ref": "issue#10/comment/c4", "weight": 1.14}]},
             {"person": "carol", "cw": 0.20, "contribution_score": 1.14,
              "signals": [{"type": "issue_comment", "ref": "issue#10/comment/c5", "weight": 1.14}]},
         ]},
        # S-104: carol owns via commits, bob co-commits, alice-arch reviews
        {"id": "S-104", "level": "Story", "job_size": 5,
         "assignees": [{"login": "carol"}],
         "contributors": [
             {"person": "carol", "cw": 0.624, "contribution_score": 8.34,
              "signals": [{"type": "commit_author", "ref": "S-104/commit/dd1", "weight": 2.78},
                          {"type": "commit_author", "ref": "S-104/commit/dd2", "weight": 2.78},
                          {"type": "commit_author", "ref": "S-104/commit/dd3", "weight": 2.78}]},
             {"person": "bob", "cw": 0.208, "contribution_score": 2.78,
              "signals": [{"type": "commit_author", "ref": "S-104/commit/dd4", "weight": 2.78}]},
             {"person": "alice-arch", "cw": 0.168, "contribution_score": 2.25,
              "signals": [{"type": "pr_reviewer", "ref": "pr#4/review/r4", "weight": 2.25}]},
         ]},
        # E-10: alice-pm leads epic via comments, bob comments
        {"id": "E-10", "level": "Epic", "job_size": 21,
         "assignees": [{"login": "alice-pm"}],
         "contributors": [
             {"person": "alice-pm", "cw": 0.75, "contribution_score": 3.42,
              "signals": [{"type": "issue_comment", "ref": "issue#e10/comment/c1", "weight": 1.14},
                          {"type": "issue_comment", "ref": "issue#e10/comment/c2", "weight": 1.14},
                          {"type": "issue_comment", "ref": "issue#e10/comment/c3", "weight": 1.14}]},
             {"person": "bob", "cw": 0.25, "contribution_score": 1.14,
              "signals": [{"type": "issue_comment", "ref": "issue#e10/comment/c4", "weight": 1.14}]},
         ]},
    ]

    return capacity, heuristics, items


def print_summary(results, iteration_id, planning_factor=0.8):
    """Print human-readable summary table."""
    print(f"\n{'='*70}")
    print(f"EDPA {VERSION} — Iteration {iteration_id}")
    print(f"{'='*70}")
    print(f"{'Person':<25} {'Role':<8} {'Capacity':>8} {'Derived':>8} {'Items':>6} {'OK':>4}")
    print(f"{'-'*70}")

    team_capacity = 0
    team_derived = 0
    all_ok = True

    for r in results:
        ok = "OK" if r["invariant_ok"] else "FAIL"
        if not r["invariant_ok"]:
            all_ok = False
        team_capacity += r["capacity"]
        team_derived += r["total_derived"]
        print(f"{r['name']:<25} {r['role']:<8} {r['capacity']:>7}h {r['total_derived']:>7}h {len(r['items']):>6} {ok:>4}")

    print(f"{'-'*70}")
    team_planning = round(team_capacity * planning_factor, 1)
    print(f"{'TEAM TOTAL':<25} {'':8} {team_capacity:>7}h {team_derived:>7}h")
    print(f"{'PLANNING CAPACITY':<25} {'':8} {team_planning:>7}h  (factor: {planning_factor})")
    print(f"\nAll invariants passed: {'YES' if all_ok else 'NO'}")

    # Per-person detail
    for r in results:
        if r["items"]:
            print(f"\n--- {r['name']} ({r['capacity']}h) ---")
            print(f"  {'Item':<10} {'Level':<8} {'JS':>4} {'CW':>6} {'Score':>7} {'Ratio':>7} {'Hours':>7}")
            for item in r["items"]:
                print(f"  {item['id']:<10} {item['level']:<8} {item['js']:>4} {item['cw']:>6.2f} {item['score']:>7.2f} {item['ratio']:>6.1%} {item['hours']:>6.1f}h")


def show_status(edpa_root):
    """Show EDPA setup status — config, team, iterations."""
    print(f"EDPA {VERSION} — Status")
    print("=" * 40)

    if not edpa_root.exists():
        print(f"\n✗ .edpa/ not found at {edpa_root}")
        print("  Run: /edpa setup \"Project Name\"")
        return

    print(f"✓ .edpa/ found at {edpa_root}")

    # People config
    people_path = edpa_root / "config" / "people.yaml"
    if people_path.exists():
        people = load_yaml(people_path) or {}
        team = people.get("people", [])
        total_fte = sum(p.get("fte", 0) for p in team)
        total_cap = sum(p.get("capacity_per_iteration", p.get("capacity", 0)) for p in team)
        print(f"✓ people.yaml — {len(team)} members, {total_fte:.1f} FTE, {total_cap:.0f}h/iteration")
        for p in team:
            cap = p.get("capacity_per_iteration", p.get("capacity", 0))
            print(f"    {p.get('name', p.get('id', '?')):<25} {p.get('role', '?'):<8} {p.get('fte', 0):.1f} FTE  {cap:.0f}h")
    else:
        print("✗ people.yaml not found")

    # Heuristics
    heuristics = load_heuristics(edpa_root)
    if heuristics:
        print("✓ heuristics loaded")
    else:
        print("✗ heuristics not found (will use defaults)")

    # Iterations — derived from .edpa/iterations/*.yaml (no longer in edpa.yaml).
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from _pi_loader import derive_pis, find_active_pi  # noqa: E402

    pis, _ = derive_pis(edpa_root)
    active_pi = find_active_pi(pis)
    iterations = active_pi.get("iterations", [])
    if iterations:
        print(f"✓ {len(iterations)} iterations defined (PI: {active_pi.get('id', '?')})")
        for it in iterations:
            status = it.get("status", "?")
            marker = "→" if status == "active" else " "
            dates = f"{it.get('start_date', '?')}–{it.get('end_date', '?')}"
            print(f"  {marker} {it.get('id', '?'):<16} {dates:<26} [{status}]")
    elif not (edpa_root / "iterations").is_dir():
        print("✗ iterations/ directory not found")

    # Backlog
    backlog_dir = edpa_root / "backlog"
    if backlog_dir.exists():
        story_count = len(list((backlog_dir / "stories").glob("*.md"))) if (backlog_dir / "stories").exists() else 0
        feature_count = len(list((backlog_dir / "features").glob("*.md"))) if (backlog_dir / "features").exists() else 0
        print(f"✓ backlog — {feature_count} features, {story_count} stories")
    else:
        print("✗ backlog/ not found")

    # Reports
    reports_dir = edpa_root / "reports"
    if reports_dir.exists():
        report_dirs = [d for d in reports_dir.iterdir() if d.is_dir()]
        if report_dirs:
            print(f"✓ {len(report_dirs)} iteration report(s)")
        else:
            print("  reports/ empty (no iterations closed yet)")

    print()


def _snapshot_payload(iteration_id, engine_output, capacity):
    """Build the snapshot dict (without frozen_at) so we can hash and
    diff successive runs without spurious revisions.
    """
    payload = {
        "snapshot_version": VERSION,
        "iteration": iteration_id,
        "generated_at": engine_output["computed_at"],
        "frozen": True,
        "methodology": engine_output["methodology"],
        # v1.14: mode field dropped — single calculation path (gates).
        "capacity_registry": {
            "people": capacity.get("people", []),
            "teams": capacity.get("teams", []),
        },
        "derived_reports": [
            {
                "person": r["id"],
                "name": r["name"],
                "role": r["role"],
                "capacity": r["capacity"],
                "total_derived": r["total_derived"],
                "items_count": len(r["items"]),
                "invariant_ok": r["invariant_ok"],
                # capacity_baseline + capacity_override are present only
                # when run_edpa applied an override — keeps the snapshot
                # byte-identical for plain runs (preserves L6 dedup).
                **({"capacity_baseline": r["capacity_baseline"]}
                   if "capacity_baseline" in r else {}),
                **({"capacity_override": r["capacity_override"]}
                   if "capacity_override" in r else {}),
            }
            for r in engine_output["people"]
        ],
        "items": [],
        "invariants": {
            "all_passed": engine_output["all_invariants_passed"],
        },
        "signature_status": "pending",
    }
    for person in engine_output["people"]:
        for item in person["items"]:
            payload["items"].append({
                "id": item["id"],
                "level": item["level"],
                "job_size": item["js"],
                "contributor": person["id"],
                "cw": item["cw"],
                "score": item["score"],
                "ratio": item["ratio"],
                "hours": item["hours"],
            })
    return payload


def _payload_signature(payload: dict) -> str:
    """Stable hash of snapshot content excluding timestamps. Two runs of
    the engine over identical inputs should hash to the same digest."""
    import hashlib
    blob = json.dumps(
        {k: v for k, v in payload.items()
         if k not in ("generated_at",)},
        sort_keys=True, ensure_ascii=False,
    )
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def write_snapshot(edpa_root, iteration_id, engine_output, capacity):
    """Write frozen snapshot to .edpa/snapshots/.

    Revisioning rule: a new _revN.json file is created only when the
    payload hash differs from the canonical PI-X.Y.json (excluding
    timestamps). Identical reruns now refresh `frozen_at` on the
    canonical file instead of proliferating PI-X.Y_rev2/3/4.json.
    """
    snapshots_dir = edpa_root / "snapshots"
    snapshots_dir.mkdir(parents=True, exist_ok=True)

    payload = _snapshot_payload(iteration_id, engine_output, capacity)
    new_signature = _payload_signature(payload)

    base = snapshots_dir / f"{iteration_id}.json"
    snapshot_path = base
    note = ""
    if base.exists():
        try:
            existing = json.load(open(base, encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            existing = None
        existing_sig = existing.get("payload_signature") if existing else None
        if existing_sig and existing_sig == new_signature:
            note = "refreshed (same content, frozen_at updated)"
        else:
            rev = 2
            while (snapshots_dir / f"{iteration_id}_rev{rev}.json").exists():
                rev += 1
            snapshot_path = snapshots_dir / f"{iteration_id}_rev{rev}.json"
            note = f"new revision (content changed); previous: {base.name}"

    payload["payload_signature"] = new_signature
    payload["frozen_at"] = datetime.now(timezone.utc).isoformat()

    with open(snapshot_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    if note:
        print(f"Snapshot {snapshot_path.name}: {note}")
    else:
        print(f"Snapshot frozen: {snapshot_path}")


def write_excel(edpa_root, iteration_id, results, capacity):
    """Write edpa-results.xlsx (Team Summary + Item Costs tabs) using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Excel export skipped (install openpyxl for XLSX output)")
        return

    report_dir = edpa_root / "reports" / f"iteration-{iteration_id}"
    report_dir.mkdir(parents=True, exist_ok=True)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def autosize(ws):
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    wb = Workbook()

    # --- Tab 1: Team Summary (per-person) ---
    ws = wb.active
    ws.title = "Team Summary"

    project_name = capacity.get("project", {}).get("name", "")
    ws.append([f"EDPA {VERSION} — {iteration_id}"])
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(bold=True, size=14)
    if project_name:
        ws.append([f"Project: {project_name}"])
        ws.merge_cells("A2:G2")
    ws.append([])

    headers = ["Person", "Role", "FTE", "Capacity (h)", "Derived (h)", "Items", "OK"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        fte = 0
        for p in capacity.get("people", []):
            if p.get("id") == r["id"]:
                fte = p.get("fte", 0)
                break
        row = [r["name"], r["role"], fte, r["capacity"],
               r["total_derived"], len(r["items"]),
               "✓" if r["invariant_ok"] else "✗"]
        ws.append(row)
        for col in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    total_cap = sum(r["capacity"] for r in results)
    total_derived = sum(r["total_derived"] for r in results)
    total_items = sum(len(r["items"]) for r in results)
    ws.append(["TOTAL", "", "", total_cap, total_derived, total_items, ""])
    for col in range(1, 8):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.border = thin_border

    autosize(ws)

    # --- Tab 2: Item Costs (per-item-person) ---
    ws2 = wb.create_sheet("Item Costs")

    ws2.append([f"EDPA {VERSION} — {iteration_id} — Per-Item Allocation"])
    ws2.merge_cells("A1:H1")
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])

    headers2 = ["Item", "Level", "JS", "Person", "CW", "Score", "Ratio", "Hours"]
    ws2.append(headers2)
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=ws2.max_row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        for item in r["items"]:
            row = [item["id"], item["level"], item["js"],
                   r["name"], item["cw"], round(item["score"], 2),
                   f"{item['ratio']:.1%}", round(item["hours"], 2)]
            ws2.append(row)
            for col in range(1, len(row) + 1):
                ws2.cell(row=ws2.max_row, column=col).border = thin_border

    autosize(ws2)

    out_path = report_dir / "edpa-results.xlsx"
    wb.save(out_path)
    print(f"Excel: {out_path}")


def main():
    parser = argparse.ArgumentParser(
        description=f"EDPA {VERSION} — Evidence-Driven Proportional Allocation Engine",
        epilog="Run with --demo to see a worked example, or --edpa-root to read from .edpa/ filesystem."
    )
    parser.add_argument("--edpa-root", help="Path to .edpa/ directory (reads backlog, config, heuristics)")
    parser.add_argument("--iteration", help="Iteration ID (e.g., PI-2026-1.3)")
    # v1.14: --mode argument removed. Engine has a single calculation
    # path (was "gates" — credits Story Done items + Feature/Epic/
    # Initiative status-transition events when git history records
    # them). When no transitions exist (e.g., backlog with no
    # sync-driven status updates), the engine degenerates to
    # Done-only credit, which is what pre-v1.14 "simple"/"full"
    # produced. So removing the modes is feature-preserving.
    parser.add_argument("--capacity", help="Path to capacity.yaml (legacy mode)")
    parser.add_argument("--heuristics", help="Path to cw_heuristics.yaml (legacy mode)")
    parser.add_argument("--output", help="Output path for edpa_results.json")
    parser.add_argument("--version", action="version", version=f"EDPA {VERSION}")
    parser.add_argument("--status", action="store_true",
                        help="Show EDPA setup status and exit")
    parser.add_argument("--demo", action="store_true",
                        help="Run with built-in sample data")
    parser.add_argument("--explain", metavar="PERSON",
                        help="Explain allocation for PERSON from already-computed results "
                             "(requires --edpa-root and --iteration). Does not re-run engine.")
    parser.add_argument("--explain-item", metavar="ITEM",
                        help="With --explain: focus on a single item ID")
    args = parser.parse_args()

    if args.explain:
        # Delegate to explain.py — read-only, no engine re-run
        sys.path.insert(0, str(Path(__file__).resolve().parent))
        try:
            from explain import explain_person, load_results  # noqa: E402
        finally:
            pass
        edpa_root = Path(args.edpa_root) if args.edpa_root else Path(".edpa")
        if not args.iteration:
            parser.error("--explain requires --iteration")
        try:
            results = load_results(edpa_root, args.iteration)
        except FileNotFoundError as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            sys.exit(1)
        md = explain_person(results, args.explain, edpa_root, args.explain_item)
        if md.startswith("ERROR:"):
            print(md, file=sys.stderr)
            sys.exit(1)
        print(md)
        sys.exit(0)

    if args.status:
        show_status(Path(args.edpa_root) if args.edpa_root else Path(".edpa"))
        sys.exit(0)

    gate_audit = None
    story_audit = None
    if args.demo:
        # Demo data has no git history — gates extraction returns 0
        # transitions and engine just credits Done items declaratively.
        print("Running EDPA demo with sample data...\n")
        capacity, heuristics, items = generate_demo_data()
        iteration_id = "DEMO-1.1"
    elif args.edpa_root:
        # Filesystem-first mode: read everything from .edpa/
        edpa_root = Path(args.edpa_root)
        if not edpa_root.exists():
            parser.error(f".edpa/ directory not found at {edpa_root}")

        capacity = load_yaml(edpa_root / "config" / "people.yaml")
        heuristics = load_heuristics(edpa_root)
        iteration_id = args.iteration

        items, manual_cw = load_backlog_items(edpa_root, iteration_id)
        # Stories carry Done credit on their own; parents (Feature/Epic/
        # Initiative) come in only as gate events synthesized from git
        # transitions. We strip Done parents from the items[] list so
        # they don't get double-counted — gate_events represent them.
        # Defects ARE credited at Done status (small bug-fix items
        # without their own gate ladder); v1.17 fix to the v1.16 silent
        # drop. Tasks behave the same way.
        items = [i for i in items if i.get("level") in ("Story", "Defect", "Task")]
        gate_events, gate_audit = load_gate_events(
            edpa_root, iteration_id, heuristics,
            people=(capacity or {}).get("people", []) or [],
        )
        items.extend(gate_events)
        # v1.17: collect yaml_edit signals from git diff over backlog/.
        # These augment existing contributors[] additively (signals stack
        # on top of detect_contributors output). When a parent Feature/
        # Epic/Initiative was seeded without contributors[], the commit
        # author who wrote the LBC + AC + NFRs gets credit automatically.
        yaml_edit_signals = {}
        try:
            sys.path.insert(0, str(Path(__file__).parent))
            from yaml_edit_signals import collect_yaml_edit_signals  # noqa: E402
            sys.path.pop(0)
            yaml_edit_signals = collect_yaml_edit_signals(
                edpa_root, iteration_id,
                heuristics.get("yaml_edit_weights"),
            )
        except Exception as exc:  # noqa: BLE001
            print(f"WARN: yaml_edit_signals collection skipped: {exc}",
                  file=sys.stderr)
        # C7.5: in-flight Story activity events — emit synthetic items
        # for Stories that have yaml_edit activity in this iteration
        # window but haven't reached Done. Credit_factor of story.js
        # (default 0.40) reserved for "ongoing work" attribution.
        story_activity_events, story_audit = load_story_activity_events(
            edpa_root, iteration_id, heuristics, yaml_edit_signals,
        )
        items.extend(story_activity_events)
        if yaml_edit_signals:
            people_for_resolve = (capacity or {}).get("people", []) or []
            _enrich_items_with_yaml_edit_signals(
                items, yaml_edit_signals, people_for_resolve,
            )
        n_yaml = sum(len(v) for v in yaml_edit_signals.values())
        done_count = (len(items) - len(gate_events)
                      - len(story_activity_events))
        print(f"Loaded {len(items)} items "
              f"({done_count} Done Stories/Defects + "
              f"{len(gate_events)} gate events + "
              f"{len(story_activity_events)} story activity events"
              f"{', ' + str(n_yaml) + ' yaml_edit signals' if n_yaml else ''})")
        if iteration_id:
            print(f"Filtered to iteration: {iteration_id}")
        if manual_cw:
            print(f"Manual CW overrides: {len(manual_cw)}")
    else:
        # Legacy mode: explicit file paths
        if not args.capacity or not args.heuristics or not args.iteration:
            parser.error("--edpa-root or (--iteration + --capacity + --heuristics) required (or --demo)")

        capacity = load_yaml(args.capacity)
        heuristics = load_yaml(args.heuristics)
        iteration_id = args.iteration
        items = []
        print(f"Legacy mode: loading from {args.capacity} and {args.heuristics}")
        print(f"NOTE: No items loaded. Use --edpa-root to read from .edpa/backlog/")

    # Resolve planning_factor from teams (team-level decision, not cadence)
    teams = capacity.get("teams", [])
    if teams:
        planning_factor = teams[0].get("planning_factor", 0.8)
    else:
        planning_factor = 0.8

    results = run_edpa(
        capacity, heuristics, items,
        edpa_root=args.edpa_root, iteration_id=iteration_id,
    )

    all_passed = all(r["invariant_ok"] for r in results if r["items"])
    team_total = sum(r["total_derived"] for r in results)

    output = {
        "iteration": iteration_id,
        "computed_at": datetime.now(timezone.utc).isoformat(),
        "methodology": f"EDPA {VERSION}",
        "planning_factor": planning_factor,
        "people": results,
        "team_total": round(team_total, 2),
        "all_invariants_passed": all_passed,
    }
    if gate_audit is not None:
        output["gate_events"] = gate_audit
    if story_audit:
        output["story_activity_events"] = story_audit

    # Write output
    if args.output:
        output_path = Path(args.output)
    elif args.edpa_root:
        output_path = Path(args.edpa_root) / "reports" / f"iteration-{iteration_id}" / "edpa_results.json"
    elif not args.demo:
        output_path = Path(f".edpa/reports/iteration-{iteration_id}/edpa_results.json")
    else:
        output_path = None

    if output_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        print(f"\nResults written to: {output_path}")

    # Write frozen snapshot
    if args.edpa_root and not args.demo:
        edpa_root = Path(args.edpa_root)
        write_snapshot(edpa_root, iteration_id, output, capacity)
        write_excel(edpa_root, iteration_id, results, capacity)

    print_summary(results, iteration_id, planning_factor)

    if not all_passed:
        sys.exit(1)


if __name__ == "__main__":
    main()
